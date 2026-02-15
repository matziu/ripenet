import io

from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from apps.ipam.models import VLAN, Host, Subnet, Tunnel
from apps.projects.models import Project, Site


class ProjectExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            import openpyxl
        except ImportError:
            return HttpResponse("openpyxl not installed", status=500)

        project = Project.objects.get(pk=project_id)
        wb = openpyxl.Workbook()

        # Sites sheet
        ws = wb.active
        ws.title = "Sites"
        ws.append(["Name", "Address", "Latitude", "Longitude"])
        for site in project.sites.all():
            ws.append([site.name, site.address, str(site.latitude or ""), str(site.longitude or "")])

        # VLANs sheet
        ws2 = wb.create_sheet("VLANs")
        ws2.append(["Site", "VLAN ID", "Name", "Purpose"])
        for vlan in VLAN.objects.filter(site__project=project).select_related("site"):
            ws2.append([vlan.site.name, vlan.vlan_id, vlan.name, vlan.purpose])

        # Subnets sheet
        ws3 = wb.create_sheet("Subnets")
        ws3.append(["Site", "VLAN", "Network", "Gateway", "Description"])
        for subnet in Subnet.objects.filter(project=project).select_related("site", "vlan"):
            site_name = subnet.site.name if subnet.site else ""
            vlan_label = f"VLAN {subnet.vlan.vlan_id}" if subnet.vlan else "(standalone)"
            ws3.append([
                site_name, vlan_label,
                str(subnet.network), str(subnet.gateway or ""), subnet.description,
            ])

        # Hosts sheet
        ws4 = wb.create_sheet("Hosts")
        ws4.append(["Site", "VLAN", "Subnet", "IP", "Hostname", "MAC", "Device Type"])
        for host in Host.objects.filter(
            subnet__project=project
        ).select_related("subnet__site", "subnet__vlan"):
            site_name = host.subnet.site.name if host.subnet.site else ""
            vlan_label = f"VLAN {host.subnet.vlan.vlan_id}" if host.subnet.vlan else "(standalone)"
            ws4.append([
                site_name,
                vlan_label,
                str(host.subnet.network),
                str(host.ip_address),
                host.hostname,
                host.mac_address,
                host.device_type,
            ])

        # Tunnels sheet
        ws5 = wb.create_sheet("Tunnels")
        ws5.append(["Name", "Type", "Subnet", "Site A", "IP A", "Site B", "IP B", "Status"])
        for tunnel in Tunnel.objects.filter(project=project).select_related("site_a", "site_b"):
            ws5.append([
                tunnel.name, tunnel.tunnel_type, str(tunnel.tunnel_subnet),
                tunnel.site_a.name, str(tunnel.ip_a),
                tunnel.site_b.name, str(tunnel.ip_b),
                tunnel.status,
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{project.name}.xlsx"'
        return response


class ProjectPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        project = Project.objects.get(pk=project_id)
        sites = project.sites.all()

        html = f"""
        <html>
        <head><style>
            body {{ font-family: sans-serif; font-size: 12px; }}
            h1 {{ color: #1a1a2e; }}
            table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 6px 8px; text-align: left; }}
            th {{ background-color: #f0f0f0; }}
        </style></head>
        <body>
        <h1>{project.name}</h1>
        <p>{project.description}</p>
        <p>Status: {project.status} | Supernet: {project.supernet or 'N/A'}</p>
        """

        for site in sites:
            html += f"<h2>{site.name}</h2><p>{site.address}</p>"
            vlans = site.vlans.prefetch_related("subnets__hosts")
            for vlan in vlans:
                html += f"<h3>VLAN {vlan.vlan_id} - {vlan.name}</h3>"
                for subnet in vlan.subnets.all():
                    html += f"<h4>{subnet.network}</h4>"
                    html += "<table><tr><th>IP</th><th>Hostname</th><th>Type</th></tr>"
                    for host in subnet.hosts.all():
                        html += f"<tr><td>{host.ip_address}</td><td>{host.hostname}</td><td>{host.device_type}</td></tr>"
                    html += "</table>"

            # Standalone subnets for this site
            standalone = Subnet.objects.filter(site=site, vlan__isnull=True).prefetch_related("hosts")
            if standalone.exists():
                html += "<h3>Standalone Subnets</h3>"
                for subnet in standalone:
                    html += f"<h4>{subnet.network}</h4>"
                    html += "<table><tr><th>IP</th><th>Hostname</th><th>Type</th></tr>"
                    for host in subnet.hosts.all():
                        html += f"<tr><td>{host.ip_address}</td><td>{host.hostname}</td><td>{host.device_type}</td></tr>"
                    html += "</table>"

        # Project-wide standalone subnets
        project_wide = Subnet.objects.filter(project=project, site__isnull=True, vlan__isnull=True).prefetch_related("hosts")
        if project_wide.exists():
            html += "<h2>Project-Wide Subnets</h2>"
            for subnet in project_wide:
                html += f"<h4>{subnet.network}</h4><p>{subnet.description}</p>"
                html += "<table><tr><th>IP</th><th>Hostname</th><th>Type</th></tr>"
                for host in subnet.hosts.all():
                    html += f"<tr><td>{host.ip_address}</td><td>{host.hostname}</td><td>{host.device_type}</td></tr>"
                html += "</table>"

        html += "</body></html>"

        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{project.name}.pdf"'
            return response
        except ImportError:
            return HttpResponse(html, content_type="text/html")
